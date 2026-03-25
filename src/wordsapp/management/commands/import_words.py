"""Import words from an Excel workbook."""

from __future__ import annotations

from argparse import ArgumentParser
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook  # type: ignore[import-untyped]

from wordsapp.models import PartOfSpeech, Record, User, Word

EXPECTED_HEADERS = ["frequency", "word", "pos", "word_ru", "word_fr"]
DEFAULT_WORKBOOK_NAME = "words.xlsx"
POS_NAME_BY_ABBREVIATION = {
    "v": "verb",
    "n": "noun",
    "r": "adverb",
    "j": "adjective",
    "u": "interjection",
    "pro": "pronoun",
    "pre": "preposition",
    "c": "conjunction",
    "d": "determiner",
}


@dataclass(frozen=True)
class PendingWordImport:
    """Validated word data pending import."""

    row_number: int
    frequency: int
    en: str
    ru: str
    fr: str
    part_of_speech: PartOfSpeech


class Command(BaseCommand):
    """Import words from an Excel workbook."""

    help = "Import Word and Record entries from an .xlsx workbook."

    def add_arguments(self, parser: ArgumentParser) -> None:
        """Add command arguments."""
        parser.add_argument(
            "--file",
            dest="workbook_path",
            default=str(Path(settings.BASE_DIR) / DEFAULT_WORKBOOK_NAME),
            help="Path to the .xlsx workbook to import.",
        )

    def handle(self, *args: object, **options: object) -> None:
        """Run the import."""
        workbook_path_option = options["workbook_path"]
        if not isinstance(workbook_path_option, str):
            raise CommandError("Workbook path must be a string.")

        workbook_path = Path(workbook_path_option).expanduser().resolve()
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")
        if workbook_path.suffix.lower() != ".xlsx":
            raise CommandError(f"Workbook must be an .xlsx file: {workbook_path}")

        try:
            user = User.objects.get(pk=1)
        except User.DoesNotExist as exc:
            raise CommandError("User with id=1 does not exist.") from exc

        part_of_speech_by_name = {
            part_of_speech.name.strip().lower(): part_of_speech
            for part_of_speech in PartOfSpeech.objects.all()
        }

        with transaction.atomic():
            pending_words = self._load_pending_words(
                workbook_path,
                part_of_speech_by_name,
            )
            self._validate_conflicts(pending_words)
            for pending_word in pending_words:
                word = Word.objects.create(
                    user_added=user,
                    en=pending_word.en,
                    ru=pending_word.ru,
                    fr=pending_word.fr,
                    part_of_speech=pending_word.part_of_speech,
                    frequency=pending_word.frequency,
                )
                Record.objects.create(word=word, user=user)

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {len(pending_words)} words from {workbook_path}."
            )
        )

    def _load_pending_words(
        self,
        workbook_path: Path,
        part_of_speech_by_name: dict[str, PartOfSpeech],
    ) -> list[PendingWordImport]:
        """Load and validate workbook rows before import."""
        workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)

            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise CommandError("Workbook is empty.") from exc

            headers = self._normalize_row(header_row)
            if headers != EXPECTED_HEADERS:
                raise CommandError(
                    "Workbook headers must be exactly: "
                    + ", ".join(EXPECTED_HEADERS)
                )

            pending_words: list[PendingWordImport] = []
            seen_frequencies: set[int] = set()
            seen_word_keys: set[tuple[str, int]] = set()

            for row_number, raw_row in enumerate(rows, start=2):
                row = self._normalize_row(raw_row)
                if not any(row):
                    continue
                if len(row) > len(EXPECTED_HEADERS):
                    raise CommandError(
                        f"Row {row_number}: expected {len(EXPECTED_HEADERS)} columns."
                    )
                if len(row) < len(EXPECTED_HEADERS):
                    row += [""] * (len(EXPECTED_HEADERS) - len(row))

                frequency_value, en, pos_value, ru, fr = row
                if not frequency_value:
                    raise CommandError(f"Row {row_number}: missing frequency.")
                if not en:
                    raise CommandError(f"Row {row_number}: missing word.")
                if not pos_value:
                    raise CommandError(f"Row {row_number}: missing pos.")

                frequency = self._parse_frequency(frequency_value, row_number)

                mapped_pos_name = POS_NAME_BY_ABBREVIATION.get(pos_value.lower())
                if mapped_pos_name is None:
                    raise CommandError(
                        f"Row {row_number}: unknown pos abbreviation '{pos_value}'."
                    )

                part_of_speech = part_of_speech_by_name.get(mapped_pos_name)
                if part_of_speech is None:
                    raise CommandError(
                        f"Row {row_number}: missing PartOfSpeech named '{mapped_pos_name}'."
                    )

                if frequency in seen_frequencies:
                    raise CommandError(
                        f"Row {row_number}: duplicate frequency {frequency} in workbook."
                    )
                seen_frequencies.add(frequency)

                word_key = (en, part_of_speech.pk)
                if word_key in seen_word_keys:
                    raise CommandError(
                        f"Row {row_number}: duplicate word '{en}' for part of speech '{part_of_speech.name}' in workbook."
                    )
                seen_word_keys.add(word_key)

                pending_words.append(
                    PendingWordImport(
                        row_number=row_number,
                        frequency=frequency,
                        en=en,
                        ru=ru,
                        fr=fr,
                        part_of_speech=part_of_speech,
                    )
                )

            return pending_words
        finally:
            workbook.close()

    def _validate_conflicts(self, pending_words: list[PendingWordImport]) -> None:
        """Ensure workbook rows do not conflict with existing data."""
        if not pending_words:
            return

        frequencies = [pending_word.frequency for pending_word in pending_words]
        existing_frequencies = {
            frequency
            for frequency in Word.objects.filter(frequency__in=frequencies).values_list(
                "frequency",
                flat=True,
            )
            if frequency is not None
        }
        if existing_frequencies:
            duplicate_frequency = min(existing_frequencies)
            raise CommandError(
                f"Import conflict: frequency {duplicate_frequency} already exists."
            )

        imported_words = {pending_word.en for pending_word in pending_words}
        imported_part_of_speech_ids = {
            pending_word.part_of_speech.pk for pending_word in pending_words
        }
        existing_word_keys = set(
            Word.objects.filter(
                en__in=imported_words,
                part_of_speech_id__in=imported_part_of_speech_ids,
            ).values_list("en", "part_of_speech_id")
        )

        for pending_word in pending_words:
            word_key = (pending_word.en, pending_word.part_of_speech.pk)
            if word_key in existing_word_keys:
                raise CommandError(
                    "Import conflict: word "
                    f"'{pending_word.en}' with part of speech "
                    f"'{pending_word.part_of_speech.name}' already exists."
                )

    @staticmethod
    def _normalize_row(raw_row: tuple[object, ...]) -> list[str]:
        """Normalize row values and drop trailing empty columns."""
        row = [Command._normalize_cell(value) for value in raw_row]
        while row and not row[-1]:
            row.pop()
        return row

    @staticmethod
    def _normalize_cell(value: object) -> str:
        """Normalize an Excel cell value into a trimmed string."""
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, bool):
            return str(value).strip()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _parse_frequency(value: str, row_number: int) -> int:
        """Parse a positive integer frequency."""
        try:
            decimal_value = Decimal(value)
        except InvalidOperation as exc:
            raise CommandError(
                f"Row {row_number}: invalid frequency '{value}'."
            ) from exc

        if decimal_value != decimal_value.to_integral_value():
            raise CommandError(f"Row {row_number}: invalid frequency '{value}'.")

        frequency = int(decimal_value)
        if frequency <= 0:
            raise CommandError(f"Row {row_number}: frequency must be positive.")
        return frequency
